"""
Compare read performance with fastexcel, xlrd and different openpyxl options
"""

from openpyxl import load_workbook
import pytest
from xlrd import open_workbook
import fastexcel


@pytest.fixture
def plain_data_xls():
    return "./python/tests/benchmarks/fixtures/plain_data.xls"

@pytest.fixture
def plain_data_xlsx():
    return "./python/tests/benchmarks/fixtures/plain_data.xlsx"

@pytest.fixture
def formula_xlsx():
    return "./python/tests/benchmarks/fixtures/plain_data.xlsx"

def pixel(test_file):
    wb = load_workbook(test_file, read_only=True, keep_links=False, data_only=True)
    for ws in wb:
        rows = ws.iter_rows()
        rows = ws.values
        for row in rows:
            for value in row:
                value


def xlread(test_file):
    wb = open_workbook(test_file)
    for ws in wb.sheets():
        for idx in range(ws.nrows):
            for value in ws.row_values(idx):
                value


def fastexcel_read(test_file):
    reader = fastexcel.read_excel(test_file)
    for sheet_name in reader.sheet_names:
        try:
            sheet = reader.load_sheet_by_name(sheet_name)
            sheet.to_arrow()
        except:
            pass


@pytest.mark.benchmark(group="xlsx")
def test_pixel(benchmark, plain_data_xlsx):
    benchmark(pixel, plain_data_xlsx)

@pytest.mark.benchmark(group="xls")
def test_xlrd(benchmark, plain_data_xls):
    benchmark(xlread, plain_data_xls)

@pytest.mark.benchmark(group="xls")
def test_fastexcel_xls(benchmark, plain_data_xls):
    benchmark(fastexcel_read, plain_data_xls)

@pytest.mark.benchmark(group="xlsx")
def test_fastexcel_xlsx(benchmark, plain_data_xlsx):
    benchmark(fastexcel_read, plain_data_xlsx)

@pytest.mark.benchmark(group="xlsx")
def test_pixel_with_formulas(benchmark, formula_xlsx):
    benchmark(pixel, formula_xlsx)

@pytest.mark.benchmark(group="xlsx")
def test_fastexcel_with_formulas(benchmark, formula_xlsx):
    benchmark(fastexcel_read, formula_xlsx)
