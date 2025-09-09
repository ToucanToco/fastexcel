from fastexcel import read_excel
from openpyxl import load_workbook
from xlrd import open_workbook


def pyxl_read(test_file_path: str):
    wb = load_workbook(test_file_path, read_only=True, keep_links=False, data_only=True)
    for ws in wb:
        rows = ws.iter_rows()
        rows = ws.values
        for row in rows:
            for _ in row:
                pass


def xlrd_read(test_file_path: str):
    wb = open_workbook(test_file_path)
    for ws in wb.sheets():
        for idx in range(ws.nrows):
            for _ in ws.row_values(idx):
                pass


def fastexcel_read(test_file_path: str):
    reader = read_excel(test_file_path)
    for sheet_name in reader.sheet_names:
        sheet = reader.load_sheet(sheet_name)
        sheet.to_arrow()
